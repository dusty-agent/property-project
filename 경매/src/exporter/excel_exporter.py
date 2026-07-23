from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import EXCEL_DIR
from src.models import AuctionProperty
from src.utils import safe_filename


class ExcelExporter:

    HEADERS = [
        ("court", "법원"),
        ("department", "담당계"),
        ("case_number", "사건번호"),
        ("item_number", "물건번호"),
        ("property_type", "물건종류"),
        ("title", "물건명"),
        ("address", "소재지"),
        ("road_address", "도로명주소"),
        ("building_name", "건물명"),
        ("appraisal_price", "감정가"),
        ("minimum_price", "최저매각가격"),
        ("minimum_price_rate", "최저가율(%)"),
        ("deposit_price", "입찰보증금"),
        ("failed_auction_count", "유찰횟수"),
        ("auction_date", "매각기일"),
        ("auction_time", "매각시간"),
        ("status", "진행상태"),
        ("land_area_m2", "토지면적(㎡)"),
        ("building_area_m2", "건물면적(㎡)"),
        ("land_share", "대지권"),
        ("sale_target", "매각대상"),
        ("claim_amount", "청구금액"),
        ("appraisal_date", "감정평가일"),
        ("remarks", "비고"),
        ("source_url", "법원 원문"),
        ("collected_at", "수집일시"),
    ]

    @classmethod
    def export(
        cls,
        properties: Iterable[AuctionProperty],
        filename: str,
    ) -> Path:
        items = list(properties)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "법원경매"

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3",
        )
        header_font = Font(bold=True)

        for column_index, (_, label) in enumerate(
            cls.HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=label,
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row_index, property_item in enumerate(
            items,
            start=2,
        ):
            item_dict = property_item.to_dict()

            for column_index, (field_name, _) in enumerate(
                cls.HEADERS,
                start=1,
            ):
                value = item_dict.get(field_name)

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, (_, label) in enumerate(
            cls.HEADERS,
            start=1,
        ):
            default_width = max(len(label) + 3, 12)

            if label in {"물건명", "소재지", "도로명주소", "비고"}:
                default_width = 35

            if label == "법원 원문":
                default_width = 45

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = default_width

        output_path = EXCEL_DIR / f"{safe_filename(filename)}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        return output_path